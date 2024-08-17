import os
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from src.fileio.file_handler import ProjectHandler
from datetime import datetime

date = datetime.today().strftime('%Y-%m-%d')


class ExcelExporter:
    def __init__(self):
        self.export_dir = 'exports'
        os.makedirs(self.export_dir, exist_ok=True)

    def apply_styles(self, ws):
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))

        # Apply styles to header
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Apply borders and alignment to all cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    def apply_conditional_formatting(self, ws, has_responded_column):
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        ws.conditional_formatting.add(
            f'{has_responded_column}2:{has_responded_column}{ws.max_row}',
            CellIsRule(operator='equal', formula=['"Yes"'], stopIfTrue=True, fill=green_fill)
        )

    def export_all_projects(self, start_date=None, end_date=None):
        wb = Workbook()
        ws = wb.active
        ws.title = "All Projects and Clients"

        headers = ["Project Name", "Project Description", "Client Name", "Client Description", "Client Company",
                   "LinkedIn", "Email", "Date Created", "Has Responded"]
        ws.append(headers)

        projects = ProjectHandler.load_projects_from_directory()

        for project in projects:
            project_date = project.get_date_created()
            if start_date and end_date and not (start_date <= project_date <= end_date):
                continue

            if not project.get_clients():
                ws.append([
                    project.get_name(),
                    project.get_description(),
                    "", "", "", "", "", "",
                    project.get_date_created_iso(),
                    "No"
                ])
            else:
                for client in project.get_clients():
                    row_data = [
                        project.get_name(),
                        project.get_description(),
                        client.get_name(),
                        client.get_description(),
                        client.get_company() if not client.get_anonymous() else "",
                        client.get_linkedin() if not client.get_anonymous() else "",
                        client.get_email() if not client.get_anonymous() else "",
                        client.get_date_created_iso(),
                        "Yes" if client.get_has_responded() else "No"
                    ]
                    ws.append(row_data)

        self.apply_styles(ws)
        self.apply_conditional_formatting(ws, 'I')  # 'I' is the column letter for "Has Responded"

        filepath = os.path.join(self.export_dir,
                                f"Full Report, ({start_date.isoformat()} - {end_date.isoformat()}).xlsx")
        wb.save(filepath)
        return filepath

    def export_specific_project(self, project_name, start_date=None, end_date=None):
        project = ProjectHandler.load_project(project_name)
        if not project:
            return None

        project_date = project.get_date_created()
        if start_date and end_date and not (start_date <= project_date <= end_date):
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = f"Project {project.get_name()}"

        headers = ["Client Name", "Client Description", "Client Company", "LinkedIn", "Email", "Date Created",
                   "Has Responded"]
        ws.append(headers)

        for client in project.get_clients():
            row_data = [
                client.get_name(),
                client.get_description(),
                client.get_company() if not client.get_anonymous() else "",
                client.get_linkedin() if not client.get_anonymous() else "",
                client.get_email() if not client.get_anonymous() else "",
                client.get_date_created_iso(),
                "Yes" if client.get_has_responded() else "No"
            ]
            ws.append(row_data)

        self.apply_styles(ws)
        self.apply_conditional_formatting(ws, 'G')  # 'G' is the column letter for "Has Responded"

        filepath = os.path.join(self.export_dir,
                                f"{project.get_name()} Report, ({start_date.isoformat()} - {end_date.isoformat()}).xlsx")
        wb.save(filepath)
        return filepath

    def export_specific_client(self, project_name, client_name, start_date=None, end_date=None):
        project = ProjectHandler.load_project(project_name)
        if not project:
            return None

        project_date = project.get_date_created()
        if start_date and end_date and not (start_date <= project_date <= end_date):
            return None

        client = next((c for c in project.get_clients() if c.get_name() == client_name), None)
        if not client:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = f"Client {client.get_name()}"

        headers = ["Client Name", "Client Description", "Client Company", "LinkedIn", "Email", "Date Created",
                   "Has Responded"]
        ws.append(headers)

        row_data = [
            client.get_name(),
            client.get_description(),
            client.get_company() if not client.get_anonymous() else "",
            client.get_linkedin() if not client.get_anonymous() else "",
            client.get_email() if not client.get_anonymous() else "",
            client.get_date_created_iso(),
            "Yes" if client.get_has_responded() else "No"
        ]
        ws.append(row_data)

        self.apply_styles(ws)
        self.apply_conditional_formatting(ws, 'G')  # 'G' is the column letter for "Has Responded"

        filepath = os.path.join(self.export_dir,
                                f"{client.get_name()} Report, ({start_date.isoformat()} - {end_date.isoformat()}).xlsx")
        wb.save(filepath)
        return filepath


class CSVExporter:
    def __init__(self):
        self.export_dir = 'exports'
        os.makedirs(self.export_dir, exist_ok=True)

    def export_all_projects(self, start_date=None, end_date=None):
        filepath = os.path.join(self.export_dir,
                                f"Full Report, ({start_date.isoformat()} - {end_date.isoformat()}).csv")

        projects = ProjectHandler.load_projects_from_directory()

        with open(filepath, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile, quoting=csv.QUOTE_ALL)
            writer.writerow(
                ["Project Name", "Project Description", "Client Name", "Client Description", "Client Company",
                 "LinkedIn", "Email", "Date Created", "Has Responded"])

            for project in projects:
                project_date = project.get_date_created()
                if start_date and end_date and not (start_date <= project_date <= end_date):
                    continue

                if not project.get_clients():
                    writer.writerow([
                        project.get_name(),
                        project.get_description().replace('\n', ' '),
                        "", "", "", "", "", "",
                        project.get_date_created_iso(),
                        ""
                    ])
                else:
                    for client in project.get_clients():
                        writer.writerow([
                            project.get_name(),
                            project.get_description().replace('\n', ' '),
                            client.get_name(),
                            client.get_description().replace('\n', ' '),
                            client.get_company() if not client.get_anonymous() else "",
                            client.get_linkedin() if not client.get_anonymous() else "",
                            client.get_email() if not client.get_anonymous() else "",
                            client.get_date_created_iso(),
                            "Yes" if client.get_has_responded() else "No"
                        ])

        return filepath

    def export_specific_project(self, project_name, start_date=None, end_date=None):
        project = ProjectHandler.load_project(project_name)
        if not project:
            return None

        project_date = project.get_date_created()
        if start_date and end_date and not (start_date <= project_date <= end_date):
            return None

        filepath = os.path.join(self.export_dir,
                                f"{project.get_name()} Report, ({start_date.isoformat()} - {end_date.isoformat()}).csv")

        with open(filepath, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile, quoting=csv.QUOTE_ALL)
            writer.writerow(
                ["Client Name", "Client Description", "Client Company", "LinkedIn", "Email", "Date Created",
                 "Has Responded"])

            for client in project.get_clients():
                writer.writerow([
                    client.get_name(),
                    client.get_description().replace('\n', ' '),
                    client.get_company() if not client.get_anonymous() else "",
                    client.get_linkedin() if not client.get_anonymous() else "",
                    client.get_email() if not client.get_anonymous() else "",
                    client.get_date_created_iso(),
                    "Yes" if client.get_has_responded() else "No"
                ])

        return filepath

    def export_specific_client(self, project_name, client_name, start_date=None, end_date=None):
        project = ProjectHandler.load_project(project_name)
        if not project:
            return None

        project_date = project.get_date_created()
        if start_date and end_date and not (start_date <= project_date <= end_date):
            return None

        client = next((c for c in project.get_clients() if c.get_name() == client_name), None)
        if not client:
            return None

        filepath = os.path.join(self.export_dir,
                                f"{client.get_name()} Report, ({start_date.isoformat()} - {end_date.isoformat()}).csv")

        with open(filepath, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile, quoting=csv.QUOTE_ALL)
            writer.writerow(
                ["Client Name", "Client Description", "Client Company", "LinkedIn", "Email", "Date Created",
                 "Has Responded"])

            writer.writerow([
                client.get_name(),
                client.get_description().replace('\n', ' '),
                client.get_company() if not client.get_anonymous() else "",
                client.get_linkedin() if not client.get_anonymous() else "",
                client.get_email() if not client.get_anonymous() else "",
                client.get_date_created_iso(),
                "Yes" if client.get_has_responded() else "No"
            ])

        return filepath
