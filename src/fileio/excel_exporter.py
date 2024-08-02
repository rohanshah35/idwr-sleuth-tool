import os
from openpyxl import Workbook
from src.fileio.file_handler import JobHandler
from src.structures.job import Job
from src.structures.client import Client


class ExcelExporter:
    def __init__(self):
        self.export_dir = 'exports'
        os.makedirs(self.export_dir, exist_ok=True)

    def export_all_jobs(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "All Jobs and Clients"

        headers = ["Job Name", "Job Description", "Client Name", "Client Description", "LinkedIn", "Email"]
        ws.append(headers)

        jobs = JobHandler.load_jobs_from_directory()

        for job in jobs:
            for client in job.get_clients():
                ws.append([
                    job.get_name(),
                    job.get_description(),
                    client.get_name(),
                    client.description,
                    client.linkedin,
                    client.get_email()
                ])

        filepath = os.path.join(self.export_dir, "all_jobs_and_clients.xlsx")
        wb.save(filepath)
        return filepath

    def export_specific_job(self, job_name):
        job = JobHandler.load_job(job_name)
        if not job:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = f"Job {job.get_name()}"

        headers = ["Client Name", "Client Description", "LinkedIn", "Email"]
        ws.append(headers)

        for client in job.get_clients():
            ws.append([
                client.get_name(),
                client.description,
                client.linkedin,
                client.get_email()
            ])

        filepath = os.path.join(self.export_dir, f"{job.get_name()}_clients.xlsx")
        wb.save(filepath)
        return filepath

    def export_specific_client(self, job_name, client_name):
        job = JobHandler.load_job(job_name)
        if not job:
            return None

        client = next((c for c in job.get_clients() if c.get_name() == client_name), None)
        if not client:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = f"Client {client.get_name()}"

        headers = ["Client Name", "Client Description", "LinkedIn", "Email"]
        ws.append(headers)

        ws.append([
            client.get_name(),
            client.description,
            client.linkedin,
            client.get_email()
        ])

        filepath = os.path.join(self.export_dir, f"{job.get_name()}_{client.get_name()}.xlsx")
        wb.save(filepath)
        return filepath


# More substantial test
def create_test_data():
    # Create jobs
    job1 = Job("Software Developer", "Developing web applications")
    job2 = Job("Graphic Designer", "Creating visual content for marketing")

    # Create clients for Software Developer job
    client1 = Client("John Doe", "Frontend specialist", "linkedin.com/in/johndoe", "john@example.com")
    client2 = Client("Jane Smith", "Backend expert", "linkedin.com/in/janesmith", "jane@example.com")
    job1.add_client(client1)
    job1.add_client(client2)

    # Create clients for Graphic Designer job
    client3 = Client("Alice Johnson", "Logo designer", "linkedin.com/in/alicejohnson", "alice@example.com")
    client4 = Client("Bob Williams", "UI/UX specialist", "", "bob@example.com")  # No LinkedIn
    client5 = Client("Charlie Brown", "Illustrator", "linkedin.com/in/charliebrown", "")  # No email
    job2.add_client(client3)
    job2.add_client(client4)
    job2.add_client(client5)

    # Save jobs
    for job in [job1, job2]:
        job_handler = JobHandler(job)
        job_handler.write_job()

    return [job1, job2]


if __name__ == "__main__":
    # Create test data
    test_jobs = create_test_data()

    exporter = ExcelExporter()

    # Export all jobs and clients
    all_jobs_file = exporter.export_all_jobs()
    print(f"All jobs and clients exported to: {all_jobs_file}")

    # Export a specific job
    job_file = exporter.export_specific_job("Software Developer")
    if job_file:
        print(f"Software Developer job exported to: {job_file}")
    else:
        print("Software Developer job not found")

    # Export a specific client
    client_file = exporter.export_specific_client("Graphic Designer", "Alice Johnson")
    if client_file:
        print(f"Client Alice Johnson exported to: {client_file}")
    else:
        print("Client Alice Johnson not found")

    # Try to export a non-existent job
    non_existent_job_file = exporter.export_specific_job("Data Scientist")
    if non_existent_job_file:
        print(f"Data Scientist job exported to: {non_existent_job_file}")
    else:
        print("Data Scientist job not found, as expected")

    # Try to export a non-existent client
    non_existent_client_file = exporter.export_specific_client("Software Developer", "Non Existent")
    if non_existent_client_file:
        print(f"Non Existent client exported to: {non_existent_client_file}")
    else:
        print("Non Existent client not found, as expected")
